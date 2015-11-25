import logging
from collections import defaultdict

from custom_user.models import AbstractEmailUser
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from django.db import models
from openpyxl import load_workbook
from xlrd import open_workbook

from locations.models import Facility, District, IP, WareHouse

logger = logging.getLogger(__name__)
PATIENTS_ADULT = "PATIENTS (ADULT)"
PATIENTS_PAED = "PATIENTS (PAED)"
CONSUMPTION = "CONSUMPTION"
LOCATION = "Facility Index"


class DashboardUser(AbstractEmailUser):
    def get_full_name(self):
        return self.email

    def get_short_name(self):
        return self.email


class FacilityCycleRecord(models.Model):
    facility = models.ForeignKey(Facility)
    cycle = models.CharField(max_length=256)
    reporting_status = models.BooleanField(default=False)
    web_based = models.BooleanField(default=False)
    multiple = models.BooleanField(default=False)

    def __unicode__(self):
        return "%s %s" % (self.facility, self.cycle)


class FacilityConsumptionRecord(models.Model):
    facility_cycle = models.ForeignKey(FacilityCycleRecord)
    opening_balance = models.FloatField(null=True, blank=True)
    quantity_received = models.FloatField(null=True, blank=True)
    pmtct_consumption = models.FloatField(null=True, blank=True)
    art_consumption = models.FloatField(null=True, blank=True)
    loses_adjustments = models.FloatField(null=True, blank=True)
    closing_balance = models.FloatField(null=True, blank=True)
    months_of_stock_of_hand = models.FloatField(null=True, blank=True)
    quantity_required_for_current_patients = models.FloatField(null=True, blank=True)
    estimated_number_of_new_patients = models.FloatField(null=True, blank=True)
    estimated_number_of_new_pregnant_women = models.FloatField(null=True, blank=True)
    packs_ordered = models.FloatField(null=True, blank=True)
    total_quantity_to_be_ordered = models.FloatField(null=True, blank=True)
    notes = models.CharField(max_length=256, null=True, blank=True)
    formulation = models.CharField(max_length=256, null=True, blank=True)

    def __unicode__(self):
        return "%s %s" % (self.facility_cycle, self.formulation)


class AdultPatientsRecord(models.Model):
    facility_cycle = models.ForeignKey(FacilityCycleRecord)
    existing = models.FloatField(null=True, blank=True)
    new = models.FloatField(null=True, blank=True)
    formulation = models.CharField(max_length=256, null=True, blank=True)

    def __unicode__(self):
        return "%s %s" % (self.facility_cycle, self.formulation)


class PAEDPatientsRecord(models.Model):
    facility_cycle = models.ForeignKey(FacilityCycleRecord)
    existing = models.FloatField(null=True, blank=True)
    new = models.FloatField(null=True, blank=True)
    formulation = models.CharField(max_length=256, null=True, blank=True)

    def __unicode__(self):
        return "%s %s" % (self.facility_cycle, self.formulation)


class WaosStandardReport():
    def __init__(self, path):
        self.path = path
        self.worksheet = self.load_worksheet()

    def load_worksheet(self):
        workbook = open_workbook(self.path)
        return workbook.sheet_by_index(0)

    def get_data(self):
        record = self.get_facility_record()
        if record:
            for n in range(4, 14):
                self.build_consumption_record(n, record)

            for n in range(16, 24):
                self.build_consumption_record(n, record)

            return record

    def build_consumption_record(self, n, record):
        formulation_name = self.worksheet.cell_value(n, 1)
        consumption_record, created = FacilityConsumptionRecord.objects.get_or_create(facility_cycle=record, formulation=formulation_name)
        consumption_record.opening_balance = self.get_numerical_value(n, 3)
        consumption_record.quantity_received = self.get_numerical_value(n, 4)
        consumption_record.pmtct_consumption = self.get_numerical_value(n, 5)
        consumption_record.art_consumption = self.get_numerical_value(n, 6)
        consumption_record.loses_adjustments = self.get_numerical_value(n, 7)
        consumption_record.closing_balance = self.get_numerical_value(n, 8)
        consumption_record.months_of_stock_of_hand = self.get_numerical_value(n, 9)
        consumption_record.quantity_required_for_current_patients = self.get_numerical_value(n, 10)
        consumption_record.estimated_number_of_new_patients = self.get_numerical_value(n, 11)
        consumption_record.estimated_number_of_new_pregnant_women = self.get_numerical_value(n, 12)
        consumption_record.total_quantity_to_be_ordered = self.get_numerical_value(n, 13)
        consumption_record.notes = self.worksheet.cell_value(n, 14)
        consumption_record.save()

    def get_numerical_value(self, n, i):
        value = self.worksheet.cell_value(n, i)
        if value:
            return value

    def get_facility_record(self):
        facility_name = self.worksheet.cell_value(27, 1)
        level = self.worksheet.cell_value(28, 1)
        full_name = "%s %s" % (facility_name, level)
        cycle = self.worksheet.cell_value(30, 1)
        try:
            location = Facility.objects.get(name__icontains=full_name)
            record, exists = FacilityCycleRecord.objects.get_or_create(facility=location, cycle=cycle)
            return record
        except ObjectDoesNotExist:
            return None


class GeneralReport():
    def __init__(self, path, cycle):
        self.path = path
        self.cycle = cycle
        self.workbook = self.get_workbook()
        self.districts = dict()
        self.ips = dict()
        self.warehouses = dict()

    def get_workbook(self):
        return load_workbook(self.path)

    def get_facility_record(self, name):
        try:
            location = Facility.objects.get(name__icontains=name)
            record, exists = FacilityCycleRecord.objects.get_or_create(facility=location, cycle=self.cycle)
            return record
        except ObjectDoesNotExist:
            return None
        except MultipleObjectsReturned:
            logger.debug("%s matched several places" % name)
            location = Facility.objects.filter(name__icontains=name)[0]
            record, exists = FacilityCycleRecord.objects.get_or_create(facility=location, cycle=self.cycle)
            return record

    def get_value(self, row, i):
        if i <= len(row):
            value = row[i].value
            if value and value != "-":
                return row[i].value

    def get_data(self):
        self.locations()
        self.adult_patients()
        self.paed_patients()
        self.consumption_records()

    def paed_patients(self):
        paed_patients_sheet = self.workbook.get_sheet_by_name(PATIENTS_PAED)
        records = defaultdict(list)
        for row in paed_patients_sheet.iter_rows('A%s:M%s' % (paed_patients_sheet.min_row + 1, paed_patients_sheet.max_row)):
            facility_name = row[1].value
            if facility_name:
                patient_record = PAEDPatientsRecord()
                patient_record.formulation = row[2].value
                patient_record.existing = self.get_value(row, 4)
                patient_record.new = self.get_value(row, 5)
                records[facility_name].append(patient_record)
            else:
                logger.debug("%s not found" % facility_name)
        for name, values in records.iteritems():
            facility_record = self.get_facility_record(name)
            patient_records = []
            if facility_record:
                PAEDPatientsRecord.objects.filter(facility_cycle=facility_record).delete()
                for r in values:
                    r.facility_cycle = facility_record
                    patient_records.append(r)
                PAEDPatientsRecord.objects.bulk_create(patient_records)

    def adult_patients(self):
        adult_patients_sheet = self.workbook.get_sheet_by_name(PATIENTS_ADULT)
        records = defaultdict(list)
        for row in adult_patients_sheet.iter_rows('A%s:M%s' % (adult_patients_sheet.min_row + 1, adult_patients_sheet.max_row)):
            facility_name = row[1].value
            if facility_name:
                patient_record = AdultPatientsRecord()
                patient_record.formulation = row[2].value
                patient_record.existing = self.get_value(row, 4)
                patient_record.new = self.get_value(row, 5)
                records[facility_name].append(patient_record)
            else:
                logger.debug("%s not found" % facility_name)
        for name, values in records.iteritems():
            facility_record = self.get_facility_record(name)
            patient_records = []
            if facility_record:
                AdultPatientsRecord.objects.filter(facility_cycle=facility_record).delete()
                for r in values:
                    r.facility_cycle = facility_record
                    patient_records.append(r)
                AdultPatientsRecord.objects.bulk_create(patient_records)

    def locations(self):
        location_sheet = self.workbook.get_sheet_by_name(LOCATION)
        facilities = []
        facility_data = []
        for row in location_sheet.iter_rows('B%s:J%s' % (location_sheet.min_row + 3, location_sheet.max_row)):
            if row[0].value:
                facility = dict()
                facility['name'] = row[0].value
                facility['status'] = row[2].value
                facility['IP'] = row[3].value
                facility['Warehouse'] = row[4].value
                facility['District'] = row[5].value
                facility['Web/Paper'] = row[7].value
                facility['Multiple'] = row[8].value
                facility_data.append(facility)
                facilities.append(self.build_facility(facility))
        Facility.objects.bulk_create(facilities)
        for f in facility_data:
            record = self.get_facility_record(f['name'])
            record.reporting_status = f['name'].strip() == 'Reporting'
            record.web_based = f['Web/Paper'].strip() == 'Web'
            record.multiple = f['Multiple'].strip() == 'Multiple orders'
            record.save()

    def consumption_records(self):
        consumption_sheet = self.workbook.get_sheet_by_name(CONSUMPTION)
        records = defaultdict(list)
        for row in consumption_sheet.iter_rows('A%s:X%s' % (consumption_sheet.min_row + 1, consumption_sheet.max_row)):
            facility_name = row[1].value
            if facility_name:
                consumption_record = FacilityConsumptionRecord()
                consumption_record.formulation = row[2].value
                consumption_record.opening_balance = self.get_value(row, 4)
                consumption_record.quantity_received = self.get_value(row, 5)
                consumption_record.pmtct_consumption = self.get_value(row, 7)
                consumption_record.art_consumption = self.get_value(row, 6)
                consumption_record.loses_adjustments = self.get_value(row, 8)
                consumption_record.closing_balance = self.get_value(row, 9)
                consumption_record.months_of_stock_of_hand = self.get_value(row, 10)
                consumption_record.quantity_required_for_current_patients = self.get_value(row, 11)
                consumption_record.estimated_number_of_new_patients = self.get_value(row, 12)
                consumption_record.estimated_number_of_new_pregnant_women = self.get_value(row, 13)
                consumption_record.packs_ordered = self.get_value(row, 14)
                records[facility_name].append(consumption_record)

        for name, values in records.iteritems():
            facility_record = self.get_facility_record(name)
            consumption_records = []
            if facility_record:
                FacilityConsumptionRecord.objects.filter(facility_cycle=facility_record).delete()
                for r in values:
                    r.facility_cycle = facility_record
                    consumption_records.append(r)
                FacilityConsumptionRecord.objects.bulk_create(consumption_records)

    def get_district(self, name):
        if name not in self.districts:
            district, _ = District.objects.get_or_create(name=name)
            self.districts[name] = district
        return self.districts[name]

    def get_ip(self, name):
        if name not in self.ips:
            ip, _ = IP.objects.get_or_create(name=name)
            self.ips[name] = ip
        return self.ips[name]

    def get_warehouse(self, name):
        if name not in self.warehouses:
            warehouse, _ = WareHouse.objects.get_or_create(name=name)
            self.warehouses[name] = warehouse
        return self.warehouses[name]

    def build_facility(self, facility):
        return Facility(name=facility['name'], warehouse=self.get_warehouse(facility['Warehouse']), ip=self.get_ip(facility['IP']), district=self.get_district(facility['District']))
