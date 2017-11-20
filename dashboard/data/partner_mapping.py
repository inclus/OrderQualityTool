import collections
from collections import OrderedDict

from openpyxl import load_workbook


class FormattedKeyDict(collections.MutableMapping):
    def __init__(self, data=None, **kwargs):
        self._store = OrderedDict()
        if data is None:
            data = {}
        self.update(data, **kwargs)

    def __setitem__(self, key, value):
        # Use the lowercased key for lookups, but store the actual
        # key alongside the value.
        self._store[self.format_key(key)] = (key, value)

    def __getitem__(self, key):
        return self._store[self.format_key(key)][1]

    def format_key(self, key):
        return ' '.join(key.lower().split())

    def __delitem__(self, key):
        del self._store[self.format_key(key)]

    def __iter__(self):
        return (casedkey for casedkey, mappedvalue in self._store.values())

    def __len__(self):
        return len(self._store)

    def lower_items(self):
        """Like iteritems(), but with all lowercase keys."""
        return (
            (lowerkey, keyval[1])
            for (lowerkey, keyval)
            in self._store.items()
        )

    def __eq__(self, other):
        if isinstance(other, collections.Mapping):
            other = FormattedKeyDict(other)
        else:
            return NotImplemented
        # Compare insensitively
        return dict(self.lower_items()) == dict(other.lower_items())

    # Copy is required
    def copy(self):
        return FormattedKeyDict(self._store.values())

    def __repr__(self):
        return str(dict(self.items()))


def load_file(mapping_file):
    workbook = load_workbook(mapping_file, read_only=True, use_iterators=True)
    only_sheet = workbook.get_active_sheet()
    mapping = FormattedKeyDict()
    for row in only_sheet.iter_rows(
                    'A%s:B%s' % (only_sheet.min_row + 1, only_sheet.max_row)):
        mapping[row[0].value] = row[1].value
    return mapping
