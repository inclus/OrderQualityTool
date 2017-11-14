import logging
from collections import defaultdict

from openpyxl import load_workbook

from dashboard.data.entities import get_real_facility_name, ExcelDataImportRecord
from dashboard.helpers import *
from dashboard.models import Cycle

logger = logging.getLogger(__name__)


class DataImport():
    def __init__(self, raw_data, cycle):
        self.raw_data = raw_data
        self.cycle = cycle
        self.name_cache = dict()
        self.locs = []
        self.pds = defaultdict(list)
        self.ads = defaultdict(list)
        self.cs = defaultdict(list)

    def build_form_db(self, cycle):
        state = cycle.state
        self.cycle = cycle.title
        self.locs = state.get(LOCS, None)
        self.pds = state.get(PDS, None)
        self.ads = state.get(ADS, None)
        self.cs = state.get(CS, None)
        return self

    def save(self):
        cycle, created = Cycle.objects.get_or_create(title=self.cycle)
        state = {ADS: self.ads, PDS: self.pds, CS: self.cs, LOCS: self.locs}
        cycle.state = state
        cycle.save()
        return cycle

    def load(self, partner_mapping=None):
        raise NotImplementedError("you need to implement this method")


def get_value(row, i):
    if i <= len(row):
        real_value = row[i].value
        value = real_value
        if "_" in str(value):
            value = real_value = value.replace("_", "-")
        if value != "-" and value != '':
            return real_value


def build_cache_by_facility(locations):
    cache = dict()
    for location in locations:
        cache[location.facility] = location
    return cache


class ExcelDataImport(DataImport):
    def get_workbook(self):
        return load_workbook(self.raw_data, read_only=True, use_iterators=True)

    def load(self, partner_mapping=None):
        self.workbook = self.get_workbook()
        self.locs = self.locations()
        locations_cache_by_facility = build_cache_by_facility(self.locs)
        self.ads = self.adult_patients(locations_cache_by_facility)
        self.pds = self.paed_patients(locations_cache_by_facility)
        self.cs = self.consumption_records(locations_cache_by_facility)
        return self

    def paed_patients(self, cache):
        paed_patients_sheet = self.workbook.get_sheet_by_name(PATIENTS_PAED_SHEET)
        records = defaultdict(list)
        for row in paed_patients_sheet.iter_rows(
                        'A%s:M%s' % (paed_patients_sheet.min_row + 1, paed_patients_sheet.max_row)):
            facility_key = get_real_facility_name(row[1].value, row[8].value)
            location = cache.get(facility_key, None)
            if facility_key and location:
                record_data = dict()
                record_data[FORMULATION] = row[2].value
                record_data[EXISTING] = get_value(row, 4)
                record_data[NEW] = get_value(row, 5)
                patient_record = ExcelDataImportRecord(data=record_data, location=location).build_patient_record()
                records[location].append(patient_record)
            else:
                logger.debug("%s not found" % facility_key)
        return records

    def adult_patients(self, cache):
        adult_patients_sheet = self.workbook.get_sheet_by_name(PATIENTS_ADULT_SHEET)
        records = defaultdict(list)
        for row in adult_patients_sheet.iter_rows(
                        'A%s:M%s' % (adult_patients_sheet.min_row + 1, adult_patients_sheet.max_row)):
            facility_key = get_real_facility_name(row[1].value, row[8].value)
            location = cache.get(facility_key, None)
            if facility_key and location:
                record_data = dict()
                record_data[FORMULATION] = row[2].value
                record_data[EXISTING] = get_value(row, 4)
                record_data[NEW] = get_value(row, 5)
                patient_record = ExcelDataImportRecord(data=record_data, location=location).build_patient_record()
                records[location].append(patient_record)
            else:
                logger.debug("%s not found" % facility_key)
        return records

    def locations(self):
        location_sheet = self.workbook.get_sheet_by_name(LOCATION)
        facility_data = []
        for row in location_sheet.iter_rows('B%s:J%s' % (location_sheet.min_row + 3, location_sheet.max_row)):
            if row[0].value:
                data = dict()
                data[SCORES] = defaultdict(dict)
                data[STATUS] = row[2].value
                data[IP] = row[3].value
                data[WAREHOUSE] = row[4].value
                data[DISTRICT] = row[5].value
                data[MULTIPLE] = row[8].value
                data[NAME] = get_real_facility_name(row[0].value, row[5].value)
                location = ExcelDataImportRecord(data).build_location()
                facility_data.append(location)
        return facility_data

    def consumption_records(self, cache):
        consumption_sheet = self.workbook.get_sheet_by_name(CONSUMPTION_SHEET)
        records = defaultdict(list)
        for row in consumption_sheet.iter_rows('A%s:X%s' % (consumption_sheet.min_row + 1, consumption_sheet.max_row)):
            facility_key = get_real_facility_name(row[1].value, row[17].value)
            location = cache.get(facility_key, None)
            if facility_key and location:
                consumption_data = dict()
                consumption_data[FORMULATION] = row[2].value
                consumption_data[OPENING_BALANCE] = get_value(row, 4)
                consumption_data[QUANTITY_RECEIVED] = get_value(row, 5)
                consumption_data[COMBINED_CONSUMPTION] = get_value(row, 6)
                consumption_data[LOSES_ADJUSTMENTS] = get_value(row, 8)
                consumption_data[CLOSING_BALANCE] = get_value(row, 9)
                consumption_data[MONTHS_OF_STOCK_ON_HAND] = get_value(row, 10)
                consumption_data[QUANTITY_REQUIRED_FOR_CURRENT_PATIENTS] = get_value(row, 11)
                consumption_data[ESTIMATED_NUMBER_OF_NEW_ART_PATIENTS] = get_value(row, 12)
                consumption_data[ESTIMATED_NUMBER_OF_NEW_PREGNANT_WOMEN] = get_value(row, 13)
                consumption_data[PACKS_ORDERED] = get_value(row, 14)
                record = ExcelDataImportRecord(data=consumption_data, location=location).build_consumption_record()
                records[location].append(record)
        return records
