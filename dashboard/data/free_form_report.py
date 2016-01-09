import logging
from collections import defaultdict

import pydash
from openpyxl import load_workbook

from dashboard.data.utils import NEW, EXISTING, FORMULATION, LOCATION, CONSUMPTION
from dashboard.helpers import PATIENTS_PAED, PATIENTS_ADULT
from dashboard.models import Cycle

logger = logging.getLogger(__name__)


class FreeFormReport():
    def __init__(self, path, cycle):
        self.path = path
        self.cycle = cycle
        self.name_cache = dict()
        self.locs = []
        self.pds = defaultdict(list)
        self.ads = defaultdict(list)
        self.cs = defaultdict(list)

    def build_form_db(self, cycle):
        state = cycle.state
        self.cycle = cycle.title
        self.locs = state.get('locs', None)
        self.pds = state.get('pds', None)
        self.ads = state.get('ads', None)
        self.cs = state.get('cs', None)
        return self

    def save(self):
        cycle, created = Cycle.objects.get_or_create(title=self.cycle)
        state = {'ads': self.ads, 'pds': self.pds, 'cs': self.cs, 'locs': self.locs}
        cycle.state = state
        cycle.save()
        return cycle

    def get_workbook(self):
        return load_workbook(self.path)

    def get_value(self, row, i):
        if i <= len(row):
            real_value = row[i].value
            value = real_value
            if value != "-" and value != '':
                return real_value

    def load(self):
        self.workbook = self.get_workbook()
        self.locs = self.locations()
        self.ads = self.adult_patients()
        self.pds = self.paed_patients()
        self.cs = self.consumption_records()
        return self

    def paed_patients(self):
        paed_patients_sheet = self.workbook.get_sheet_by_name(PATIENTS_PAED)
        records = defaultdict(list)
        for row in paed_patients_sheet.iter_rows(
                        'A%s:M%s' % (paed_patients_sheet.min_row + 1, paed_patients_sheet.max_row)):
            facility_key, facility_name = self.get_facility_name(row)
            if facility_name:
                patient_record = dict()
                patient_record[FORMULATION] = row[2].value
                patient_record[EXISTING] = self.get_value(row, 4)
                patient_record[NEW] = self.get_value(row, 5)
                records[facility_name].append(patient_record)
            else:
                logger.debug("%s not found" % facility_name)
        return records

    def adult_patients(self):
        adult_patients_sheet = self.workbook.get_sheet_by_name(PATIENTS_ADULT)
        records = defaultdict(list)
        for row in adult_patients_sheet.iter_rows(
                        'A%s:M%s' % (adult_patients_sheet.min_row + 1, adult_patients_sheet.max_row)):
            facility_key, facility_name = self.get_facility_name(row)
            if facility_name:
                patient_record = dict()
                patient_record[FORMULATION] = row[2].value
                patient_record[EXISTING] = self.get_value(row, 4)
                patient_record[NEW] = self.get_value(row, 5)
                records[facility_name].append(patient_record)
            else:
                logger.debug("%s not found" % facility_name)
        return records

    def locations(self):
        location_sheet = self.workbook.get_sheet_by_name(LOCATION)
        facility_data = []
        for row in location_sheet.iter_rows('B%s:J%s' % (location_sheet.min_row + 3, location_sheet.max_row)):
            if row[0].value:
                facility = dict()
                facility['scores'] = defaultdict(dict)
                facility['name'] = row[0].value
                facility['status'] = row[2].value
                facility['IP'] = row[3].value
                facility['Warehouse'] = row[4].value
                facility['District'] = row[5].value
                facility['Web/Paper'] = row[7].value
                facility['Multiple'] = row[8].value
                facility_data.append(facility)
        return facility_data

    def consumption_records(self):
        consumption_sheet = self.workbook.get_sheet_by_name(CONSUMPTION)
        records = defaultdict(list)
        for row in consumption_sheet.iter_rows('A%s:X%s' % (consumption_sheet.min_row + 1, consumption_sheet.max_row)):
            facility_key, facility_name = self.get_facility_name(row)
            if facility_name:
                consumption_record = dict()
                consumption_record[FORMULATION] = row[2].value
                consumption_record['opening_balance'] = self.get_value(row, 4)
                consumption_record['quantity_received'] = self.get_value(row, 5)
                consumption_record['pmtct_consumption'] = self.get_value(row, 7)
                consumption_record['art_consumption'] = self.get_value(row, 6)
                consumption_record['loses_adjustments'] = self.get_value(row, 8)
                consumption_record['closing_balance'] = self.get_value(row, 9)
                consumption_record['months_of_stock_of_hand'] = self.get_value(row, 10)
                consumption_record['quantity_required_for_current_patients'] = self.get_value(row, 11)
                consumption_record['estimated_number_of_new_patients'] = self.get_value(row, 12)
                consumption_record['estimated_number_of_new_pregnant_women'] = self.get_value(row, 13)
                consumption_record['packs_ordered'] = self.get_value(row, 14)
                records[facility_key].append(consumption_record)
        return records

    def get_facility_name(self, row):
        facility_name = row[1].value
        if facility_name:
            if facility_name not in self.name_cache:
                locations = pydash.chain(self.locs).reject(lambda x: x['name'] is None).select(
                        lambda x: facility_name in x['name']).value()
                if len(locations) > 0:
                    facility_key = locations[0]['name']
                else:
                    facility_key = facility_name

                self.name_cache[facility_name] = facility_key
            else:
                facility_key = self.name_cache[facility_name]
        else:
            facility_key = facility_name
        return facility_key, facility_name
