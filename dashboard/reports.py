from collections import defaultdict

from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from openpyxl import load_workbook

from dashboard.helpers import PATIENTS_ADULT, PATIENTS_PAED
from dashboard.models import Consumption, Cycle, logger, PAEDPatientsRecord, AdultPatientsRecord, LOCATION, CONSUMPTION
from locations.models import Facility, District, IP, WareHouse


class GeneralReport():
    def __init__(self, path, cycle):
        self.path = path
        self.cycle = cycle
        self.workbook = self.get_workbook()
        self.districts = dict()
        self.ips = dict()
        self.warehouses = dict()

    def get_workbook(self):
        return load_workbook(self.path)

    def get_facility_record(self, name):
        try:
            location = Facility.objects.get(name__icontains=name)
            record, exists = Cycle.objects.get_or_create(facility=location, cycle=self.cycle)
            return record
        except ObjectDoesNotExist:
            return None
        except MultipleObjectsReturned:
            logger.debug("%s matched several places" % name)
            location = Facility.objects.filter(name__icontains=name)[0]
            record, exists = Cycle.objects.get_or_create(facility=location, cycle=self.cycle)
            return record

    def get_value(self, row, i):
        if i <= len(row):
            real_value = row[i].value
            value = real_value
            if value != "-" and value != '':
                return real_value

    def get_data(self):
        self.locations()
        self.adult_patients()
        self.paed_patients()
        self.consumption_records()

    def paed_patients(self):
        paed_patients_sheet = self.workbook.get_sheet_by_name(PATIENTS_PAED)
        records = defaultdict(list)
        for row in paed_patients_sheet.iter_rows('A%s:M%s' % (paed_patients_sheet.min_row + 1, paed_patients_sheet.max_row)):
            facility_name = row[1].value
            if facility_name:
                patient_record = PAEDPatientsRecord()
                patient_record.formulation = row[2].value
                patient_record.existing = self.get_value(row, 4)
                patient_record.new = self.get_value(row, 5)
                records[facility_name].append(patient_record)
            else:
                logger.debug("%s not found" % facility_name)
        for name, values in records.items():
            facility_record = self.get_facility_record(name)
            patient_records = []
            if facility_record:
                PAEDPatientsRecord.objects.filter(facility_cycle=facility_record).delete()
                for r in values:
                    r.facility_cycle = facility_record
                    patient_records.append(r)
                PAEDPatientsRecord.objects.bulk_create(patient_records)

    def adult_patients(self):
        adult_patients_sheet = self.workbook.get_sheet_by_name(PATIENTS_ADULT)
        records = defaultdict(list)
        for row in adult_patients_sheet.iter_rows('A%s:M%s' % (adult_patients_sheet.min_row + 1, adult_patients_sheet.max_row)):
            facility_name = row[1].value
            if facility_name:
                patient_record = AdultPatientsRecord()
                patient_record.formulation = row[2].value
                patient_record.existing = self.get_value(row, 4)
                patient_record.new = self.get_value(row, 5)
                records[facility_name].append(patient_record)
            else:
                logger.debug("%s not found" % facility_name)
        for name, values in records.items():
            facility_record = self.get_facility_record(name)
            patient_records = []
            if facility_record:
                AdultPatientsRecord.objects.filter(facility_cycle=facility_record).delete()
                for r in values:
                    r.facility_cycle = facility_record
                    patient_records.append(r)
                AdultPatientsRecord.objects.bulk_create(patient_records)

    def locations(self):
        location_sheet = self.workbook.get_sheet_by_name(LOCATION)
        facilities = []
        facility_data = []
        for row in location_sheet.iter_rows('B%s:J%s' % (location_sheet.min_row + 3, location_sheet.max_row)):
            if row[0].value:
                facility = dict()
                facility['name'] = row[0].value
                facility['status'] = row[2].value
                facility['IP'] = row[3].value
                facility['Warehouse'] = row[4].value
                facility['District'] = row[5].value
                facility['Web/Paper'] = row[7].value
                facility['Multiple'] = row[8].value
                facility_data.append(facility)
                self.build_facility(facility)

        for f in facility_data:
            record = self.get_facility_record(f['name'])
            record.reporting_status = f['status'].strip() == 'Reporting'
            record.web_based = f['Web/Paper'].strip() == 'Web'
            record.multiple = f['Multiple'].strip() == 'Multiple orders'
            record.save()

    def consumption_records(self):
        consumption_sheet = self.workbook.get_sheet_by_name(CONSUMPTION)
        records = defaultdict(list)
        for row in consumption_sheet.iter_rows('A%s:X%s' % (consumption_sheet.min_row + 1, consumption_sheet.max_row)):
            facility_name = row[1].value
            if facility_name:
                consumption_record = Consumption()
                consumption_record.formulation = row[2].value
                consumption_record.opening_balance = self.get_value(row, 4)
                consumption_record.quantity_received = self.get_value(row, 5)
                consumption_record.pmtct_consumption = self.get_value(row, 7)
                consumption_record.art_consumption = self.get_value(row, 6)
                consumption_record.loses_adjustments = self.get_value(row, 8)
                consumption_record.closing_balance = self.get_value(row, 9)
                consumption_record.months_of_stock_of_hand = self.get_value(row, 10)
                consumption_record.quantity_required_for_current_patients = self.get_value(row, 11)
                consumption_record.estimated_number_of_new_patients = self.get_value(row, 12)
                consumption_record.estimated_number_of_new_pregnant_women = self.get_value(row, 13)
                consumption_record.packs_ordered = self.get_value(row, 14)
                records[facility_name].append(consumption_record)

        for name, values in records.items():
            facility_record = self.get_facility_record(name)
            consumption_records = []
            if facility_record:
                Consumption.objects.filter(facility_cycle=facility_record).delete()
                for r in values:
                    r.facility_cycle = facility_record
                    consumption_records.append(r)
                Consumption.objects.bulk_create(consumption_records)

    def get_district(self, name):
        if name not in self.districts:
            district, _ = District.objects.get_or_create(name=name)
            self.districts[name] = district
        return self.districts[name]

    def get_ip(self, name):
        if name not in self.ips:
            ip, _ = IP.objects.get_or_create(name=name)
            self.ips[name] = ip
        return self.ips[name]

    def get_warehouse(self, name):
        if name not in self.warehouses:
            warehouse, _ = WareHouse.objects.get_or_create(name=name)
            self.warehouses[name] = warehouse
        return self.warehouses[name]

    def build_facility(self, facility):
        warehouse = self.get_warehouse(facility['Warehouse'])
        ip = self.get_ip(facility['IP'])
        district = self.get_district(facility['District'])
        facility_name = facility['name']
        facility, _ = Facility.objects.get_or_create(name=facility_name)
        facility.warehouse = warehouse
        facility.ip = ip
        facility.district = district
        facility.save()
        return facility
